import pytest
import openpyxl
import os
from time import sleep
import time

import pytest
from selenium.webdriver import Chrome, ChromeOptions
from selenium.webdriver.common.by import By

o = ChromeOptions()
o.add_experimental_option('detach', True)
o.add_experimental_option("prefs", {"safebrowsing.enabled": True})
o.add_argument("--disable-notifications")

driver = Chrome(options=o)
driver.implicitly_wait(10)
driver.get("https://www.saucedemo.com/")

wb = openpyxl.Workbook()
sheetName = "sheet1"
if sheetName in wb.sheetnames:
    ws = wb[sheetName]
else:
    ws = wb.create_sheet(sheetName)
# print(wb[sheetName])
ws['A1'] = 'user'
ws['B1'] = 'password'

ws.append(['standard_user', 'secret_sauce'])
ws.append(['user1', '1234'])
ws.append(['locked_out_user', 'secret_sauce'])

wb.save("sample1.xlsx")

ans = []


def getdata():
    wb1 = openpyxl.load_workbook('sample1.xlsx')
    ws2 = wb1["sheet1"]
    data = []

    for i in ws2.iter_rows(min_row=2, values_only=True):
        data.append(i)
    return data


@pytest.mark.parametrize('usr,psw', getdata())
def test_check(usr, psw):
    driver.find_element(By.XPATH, "//input[@id='user-name']").send_keys(usr)
    driver.find_element(By.XPATH, "//input[@id='password']").send_keys(psw)
    driver.find_element(By.ID, "login-button").click()
    assert driver.current_url == 'https://www.saucedemo.com/inventory.html', driver.refresh()
    sleep(5)
    driver.back()
